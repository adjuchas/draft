import pymysql
import openpyxl


yb_mysql = pymysql.connect(
    host='127.0.0.1',
    port=3306,
    user='root',
    password='csxyyiban1228mysql',
    database='app_215',
    charset='utf8'
)
cursor = yb_mysql.cursor()

sql = 'insert into admin215 values(%s, %s);'


book = openpyxl.load_workbook('lvzhou-admin.xlsx')
sheet = book.active

row = 2
all_row = sheet.max_row

for a in range(row, all_row+1):
    stuid = sheet.cell(a, 1).value
    yb_realname = sheet.cell(a, 2).value
    cursor.execute(sql, [stuid, yb_realname])

yb_mysql.commit()
cursor.close()

print("success for into admin")
