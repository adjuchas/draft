from django.conf import settings
from .sql import AdminSql
import time
import os
import shutil
import zipfile
import random
import openpyxl


def create_name():
    t = ''
    a = list(time.localtime())
    for i in a[1: 6]:
        t += str(i)
    t += str(random.randint(10, 99)) + '.zip'
    return t


def create_url_list(down_list):
    url_1 = os.path.join(settings.MEDIA_ROOT, "yibanlvzhou")
    all_list = []
    for recode_id, recode in down_list.items():
        file_name = str(recode[0]) + "-" + str(recode[2])
        url_2 = os.path.join(url_1, str(recode[1]), file_name)
        all_list.append(url_2)
    return all_list


def create_download_url():
    down_zip = "/215_down"
    if not os.path.isdir(down_zip):
        os.mkdir(down_zip)
    down_zip = "/215_down/lvzhou_admin"
    if not os.path.isdir(down_zip):
        os.mkdir(down_zip)
    return down_zip


def create_zip(down_list):
    zip_name = create_name()
    zip_url = create_download_url()
    file_url_list = create_url_list(down_list)
    zip_file = zipfile.ZipFile(zip_name, 'w')
    for i in file_url_list:
        zip_file.write(i, os.path.basename(i), compress_type=zipfile.ZIP_DEFLATED)
    zip_file.close()
    nowurl = os.path.join(os.getcwd(), zip_name)
    shutil.copyfile(nowurl, os.path.join(zip_url, zip_name))
    os.remove(nowurl)
    return os.path.join(zip_url, zip_name)


def write_xlsx_head(sheet):
    sheet.cell(1, 1).value = "学号"
    sheet.cell(1, 2).value = "姓名"
    sheet.cell(1, 3).value = "作品名"
    sheet.cell(1, 4).value = "作品类型"
    sheet.cell(1, 5).value = "作品状态"
    sheet.column_dimensions['A'].width = 17
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 30


def write_xlsx_recode(sheet, recode_list):
    stuid_list = []
    row = 2
    for recode in recode_list:
        sheet.cell(row, 1).value = recode[0]
        stuid_list.append(recode[0])
        sheet.cell(row, 3).value = recode[1]
        sheet.cell(row, 4).value = settings.LVZHOU[str(recode[2])]
        sheet.cell(row, 5).value = recode[3]
        row += 1
    return stuid_list


def write_xlsx_name(sheet, name_list):
    row = 2
    for name in name_list:
        sheet.cell(row, 2).value = name
        row += 1


def create_xlsx(state, recode_list):
    book = openpyxl.Workbook()
    sheet = book.active
    write_xlsx_head(sheet)
    stuid_list = write_xlsx_recode(sheet, recode_list)
    adminsql = AdminSql()
    name_list = adminsql.find_names(stuid_list)
    write_xlsx_name(sheet, name_list)
    xlsx_name = state + '.xlsx'
    book.save(xlsx_name)
    return xlsx_name


def create_xlsx_zip(name_list):
    nowurl = os.getcwd()
    url_list = []
    for name in name_list:
        url = os.path.join(nowurl, name)
        url_list.append(url)
    zip_name = create_name()
    zip_url = create_download_url()
    zip_file = zipfile.ZipFile(zip_name, 'w')
    for i in url_list:
        zip_file.write(i, os.path.basename(i), compress_type=zipfile.ZIP_DEFLATED)
    zip_file.close()
    nowzip = os.path.join(nowurl, zip_name)
    shutil.copyfile(nowzip, os.path.join(zip_url, zip_name))
    os.remove(nowzip)
    return os.path.join(zip_url, zip_name)
